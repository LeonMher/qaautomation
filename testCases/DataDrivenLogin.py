from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.common.by import By


# Initialize the WebDriver (change the path to your WebDriver executable)
driver = webdriver.Chrome()

# Load the Excel file
excel_file = "C:\\Users\\inOne_02\\Desktop\\testTheTest\\qaautomation\\testCases\\login.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook["Login"]

# Iterate through rows in the Excel file
for row in sheet.iter_rows(min_row=2, values_only=True):
    username = row[0]
    password = row[1]

    # Open the login page (replace with your login page URL)
    driver.get("https://demo.nopcommerce.com/login?returnUrl=%2F")
    # Find and fill in the username and password fields
    username_field = driver.find_element(By.ID, "Email" )  # Replace with the actual element locator
    password_field = driver.find_element(By.ID, "Password")
    username_field.send_keys(username)
    password_field.send_keys(password)

    # Submit the login form (replace with the actual submit action)
    password_field.send_keys(Keys.RETURN)

    # Add any further automation steps after successful login

# Close the WebDriver when done
driver.quit()
